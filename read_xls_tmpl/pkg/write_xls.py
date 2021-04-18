from openpyxl import Workbook
import os, sys
from pathlib import Path
fileDir = os.path.dirname(os.path.realpath('__file__'))
sys.path.append(fileDir)


def print_xls(person_export, header):
    #if str_datei == "":
    #    str_xlsfile=os.path.join(fileDir, 'read_xls_tmpl', 'xls', 'export.xlsx')
    #    file_xls=Path(str_xlsfile)
    #    if file_xls.is_file():
    #        os.remove(str_xlsfile)

    str_xlsfile=os.path.join(fileDir, 'read_xls_tmpl', 'xls', 'export.xlsx')
    wb=Workbook()
    sheet = wb.create_sheet('Ergebnis', 0)
    # Header Schreiben
    for x, entry in enumerate(header):
        sheet.cell(row=1, column=x+1, value=entry)
    # Rest schreiben
    y=2
    for person in person_export:
        sheet.cell(row=y, column=2, value=person.pk)
        sheet.cell(row=y, column=3, value=person.station)
        sheet.cell(row=y, column=4, value=person.flotte)
        y=y+1
    freezy=sheet['Y2']
    sheet.freeze_panes=freezy
    sheet.auto_filter.ref = "A1:Y1"
    wb.save(str_xlsfile)

def find_column_vvt(vvt_entry):
    pass
